import pytest
import openpyxl
from constants import globalConstants
from pathlib import Path
from datetime import date
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec

# test yapacagimiz class ve method'larin onune test_ prefix'i yazmamiz sart
class Test_DemoClass:

    # her testden once cagirilir
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        
        # gunun tarihini al, bu tarih ile bir klasor var mi kontrol et, yoksa olustur
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)
        
    # her testden sonra cagirilir
    def teardown_method(self):
        self.driver.quit()

    def readData(self):
        print("x")

    # setup -> test_demoFunc -> teardown
    def test_demoFunc(self):
        # #A Act Arrange Assert
        text = "Hello"
        assert text == "Hello"

    # setup -> test_demo2 -> teardown
    def test_demo2(self):
        assert True

    def getData():
        # veriyi al
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile["sheet1"]

        totalRows = selectedSheet.max_row
        data = []
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i, 1).value
            password = selectedSheet.cell(i, 2).value
            tupleData = (username, password)
            data.append(tupleData)
        return data

    @pytest.mark.parametrize("username, password", getData())
    def test_invalid_login(self, username, password):
        self.waitForElementVisible((By.ID, "user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"), 10)
        passwordInput = self.driver.find_element(By.ID, "password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID, "login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH, "/html/body/div/div/div[2]/div[1]/div/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        # magic string ("bu sekilde metinsel olarak yazilmis string yazilar")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    def waitForElementVisible(self, locator, timeout=5):
        WebDriverWait(self.driver, timeout).until(ec.visibility_of_element_located(locator))  
