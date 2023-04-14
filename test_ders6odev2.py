import pytest
import openpyxl
from pathlib import Path
from datetime import date
from selenium import webdriver
from constants import loginConstants
from constants import errorConstants
from constants import xpathConstants
from constants import globalConstants
from constants import screenConstants
from constants import decorationConstants
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec

class Test_SauceDemo:

    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)

        self.usernameInput = self.driver.find_element(By.ID, f"{loginConstants.logUser}")
        self.passwordInput = self.driver.find_element(By.ID, f"{loginConstants.logPassword}")
        self.loginBtn = self.driver.find_element(By.ID,f"{loginConstants.loginButton}")
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)
        
    def teardown_method(self):
        self.driver.quit()

    def waitForElementVisible(self, locator, timeout=5):
        WebDriverWait(self.driver, timeout).until(ec.visibility_of_element_located(locator))  

    def readData(self):
        print("x")

    @pytest.mark.skipif(reason = errorConstants.emptyUPError)
    def test_emptyUsernameAndPassword(self):
        self.waitForElementVisible((By.ID, f"{loginConstants.logUser}"))
        self.waitForElementVisible((By.ID, f"{loginConstants.logPassword}"))
        self.loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH, f"{xpathConstants.emptyUPPath}")

        if self.usernameInput.text == " " and self.passwordInput.text == " ":
            assert errorMessage.text == f"{errorConstants.emptyUPError}"
        self.driver.save_screenshot(f"{self.folderPath}/{screenConstants.emptyUPscreen}")
        
    @pytest.mark.filterwarnings(f"{decorationConstants.warning}")
    def test_emptyPassword(self):
        self.waitForElementVisible((By.ID, f"{loginConstants.logUser}"))
        self.waitForElementVisible((By.ID, f"{loginConstants.logPassword}"))
        self.usernameInput.send_keys(loginConstants.userName)
        self.loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH, f"{xpathConstants.emptyPPath}")
        
        if self.passwordInput.text == " ":
            assert errorMessage.text == f"{errorConstants.emptyPError}"
        self.driver.save_screenshot(f"{self.folderPath}/{screenConstants.emptyPscreen}")

    @pytest.fixture()
    def test_lockedOut(self):
        self.waitForElementVisible((By.ID, f"{loginConstants.logUser}"))
        self.waitForElementVisible((By.ID, f"{loginConstants.logPassword}"))
        self.usernameInput.send_keys(loginConstants.lockedUser)
        self.passwordInput.send_keys(loginConstants.lockedPassword)
        self.loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH, f"{xpathConstants.lockedOutPath}")
        assert errorMessage.text == f"{errorConstants.lockedOutError}"
        self.driver.save_screenshot(f"{self.folderPath}/{screenConstants.lockedScreen}")


    def getData():
        excelFile = openpyxl.load_workbook("data/test_wrongUser.xlsx")
        selectedSheet = excelFile["Sheet"]

        totalRows = selectedSheet.max_row
        data = []
        for i in range(2, totalRows + 1):
            username = selectedSheet.cell(i, 1).value
            password = selectedSheet.cell(i, 2).value
            tupleData = (username, password)
            data.append(tupleData)
        return data

    @pytest.mark.parametrize(f"{decorationConstants.parametrize}", getData())
    def test_wrongUser(self, username, password):
        self.waitForElementVisible((By.ID, f"{loginConstants.logUser}"))
        self.waitForElementVisible((By.ID, f"{loginConstants.logPassword}"))
        self.usernameInput.send_keys(username)
        self.passwordInput.send_keys(password)
        self.loginBtn.click()
        assert self.driver.find_element(By.XPATH, f"{xpathConstants.wrongUserPath}")
        self.driver.save_screenshot(f"{self.folderPath}/{screenConstants.wrongScreen}")

    @pytest.mark.skip()
    def test_xButton(self):
        self.waitForElementVisible((By.ID, f"{loginConstants.logUser}"))
        self.waitForElementVisible((By.ID, f"{loginConstants.logPassword}"))
        self.usernameInput.send_keys(loginConstants.xUsername)
        self.passwordInput.send_keys(loginConstants.xPassword)
        self.loginBtn.click()

        errorButton = self.driver.find_element(By.XPATH, f"{xpathConstants.xbuttonPath}")
        errorButton.click()
        self.driver.save_screenshot(f"{self.folderPath}/{screenConstants.xbuttonScreen}")
    
    def test_loggingIn(self):
        self.waitForElementVisible((By.ID, f"{loginConstants.logUser}"))
        self.waitForElementVisible((By.ID, f"{loginConstants.logPassword}"))
        self.usernameInput.send_keys(loginConstants.user)
        self.passwordInput.send_keys(loginConstants.password)
        self.loginBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/{screenConstants.loginScreen}")
        